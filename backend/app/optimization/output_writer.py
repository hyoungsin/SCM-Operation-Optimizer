from pathlib import Path

from openpyxl import Workbook

from app.core.config import OUTPUT_DIR


def write_output_workbook(run_id: str, solve_result: dict[str, object]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{run_id}_output.xlsx"
    tables = solve_result.get("tables", {})
    week_keys = []
    if tables.get("shipment_sea"):
        week_keys = [key for key in tables["shipment_sea"][0].keys() if key != "model-site"]

    workbook = Workbook()
    shipment_sea_sheet = workbook.active
    shipment_sea_sheet.title = "shipment-sea"
    shipment_sea_sheet.append(["model-site", *week_keys])
    for row in tables.get("shipment_sea", []):
        shipment_sea_sheet.append([row.get("model-site", ""), *[int(row.get(week, 0)) for week in week_keys]])

    shipment_air_sheet = workbook.create_sheet("shipment-air")
    shipment_air_sheet.append(["model-site", *week_keys])
    for row in tables.get("shipment_air", []):
        shipment_air_sheet.append([row.get("model-site", ""), *[int(row.get(week, 0)) for week in week_keys]])

    shortage_sheet = workbook.create_sheet("shortage")
    shortage_sheet.append(["model-site", *week_keys])
    for row in tables.get("shortage", []):
        shortage_sheet.append([row.get("model-site", ""), *[int(row.get(week, 0)) for week in week_keys]])

    workbook.save(output_path)
    return output_path
