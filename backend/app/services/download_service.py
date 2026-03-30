from pathlib import Path

from openpyxl import Workbook

from app.core.config import REPORT_DIR
from app.repository.run_repository import get_run, update_run


def get_validation_report_file_path(run_id: str) -> Path | None:
    run = get_run(run_id)
    if run is None:
        return None

    validation = run.get("validation")
    if not validation:
        return None

    existing_report_path = run.get("validation_report_xlsx_path")
    if isinstance(existing_report_path, Path) and existing_report_path.exists():
        return existing_report_path

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / f"{run_id}_validation_report.xlsx"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["field", "value"])
    summary_sheet.append(["run_id", run_id])
    summary_sheet.append(["validation_status", validation.get("validation_status", "")])
    summary_sheet.append(["error_count", validation.get("summary", {}).get("errors", 0)])
    summary_sheet.append(["warning_count", validation.get("summary", {}).get("warnings", 0)])
    summary_sheet.append(["info_count", validation.get("summary", {}).get("infos", 0)])
    summary_sheet.append(["solve_allowed", str(validation.get("solve_allowed", False))])

    details_sheet = workbook.create_sheet("details")
    details_sheet.append(["level", "sheet", "row", "column", "message", "action_taken"])

    for error in validation.get("errors", []):
        details_sheet.append(
            [
                "error",
                error.get("sheet", ""),
                error.get("row", ""),
                error.get("column", ""),
                error.get("message", ""),
                "Fix input data before solve",
            ]
        )

    for warning in validation.get("warnings", []):
        details_sheet.append(
            [
                "warning",
                warning.get("sheet", ""),
                warning.get("row", ""),
                warning.get("column", ""),
                warning.get("message", ""),
                "Corrected during validation",
            ]
        )

    workbook.save(report_path)
    update_run(run_id, validation_report_xlsx_path=report_path)
    return report_path


def get_output_file_path(run_id: str) -> Path | None:
    run = get_run(run_id)
    if run is None:
        return None
    output_path = run.get("output_file_path")
    if isinstance(output_path, Path):
        return output_path
    return None


def get_corrected_preview_file_path(run_id: str) -> Path | None:
    run = get_run(run_id)
    if run is None:
        return None

    preview = run.get("preview")
    if not preview:
        return None

    existing_preview_path = run.get("preview_file_path")
    if isinstance(existing_preview_path, Path) and existing_preview_path.exists():
        return existing_preview_path

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    preview_path = REPORT_DIR / f"{run_id}_corrected_preview.xlsx"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["field", "value"])
    summary_sheet.append(["run_id", run_id])
    summary_sheet.append(["preview_generated", str(preview.get("preview_generated", False))])
    summary_sheet.append(["sheet_count", len(preview.get("sheets", []))])
    summary_sheet.append(["total_row_count", sum(sheet.get("row_count", 0) for sheet in preview.get("sheets", []))])

    for sheet in preview.get("sheets", []):
        sheet_name = str(sheet.get("sheet_name", "sheet"))[:31] or "sheet"
        worksheet = workbook.create_sheet(sheet_name)
        sample_rows = sheet.get("sample_rows", [])
        headers = []
        for row in sample_rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        worksheet.append(["sheet_name", "row_count", *headers])
        for index, row in enumerate(sample_rows):
            prefix = [sheet.get("sheet_name", "") if index == 0 else "", sheet.get("row_count", "") if index == 0 else ""]
            worksheet.append([*prefix, *[row.get(header, "") for header in headers]])

    workbook.save(preview_path)
    update_run(run_id, preview_file_path=preview_path)
    return preview_path
