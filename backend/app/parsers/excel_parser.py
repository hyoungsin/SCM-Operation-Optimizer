from pathlib import Path

from openpyxl import load_workbook


def get_workbook_sheet_names(file_path: Path) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def parse_workbook(file_path: Path) -> dict[str, object]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheets: dict[str, dict[str, object]] = {}
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]] if rows else []
            data_rows: list[dict[str, object]] = []

            for row in rows[1:]:
                row_dict: dict[str, object] = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row[index] if index < len(row) else None
                if any(value is not None and str(value).strip() != "" for value in row_dict.values()):
                    data_rows.append(row_dict)

            sheets[sheet_name] = {
                "headers": headers,
                "rows": data_rows,
                "row_count": len(data_rows),
            }

        return {
            "file_path": str(file_path),
            "sheet_names": list(sheets.keys()),
            "sheet_count": len(sheets),
            "sheets": sheets,
        }
    finally:
        workbook.close()
